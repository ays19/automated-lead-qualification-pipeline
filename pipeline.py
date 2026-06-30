#!/usr/bin/env python3
"""
Automated Lead Qualification Pipeline
======================================
Ingests LinkedIn profile URLs from a CSV, scores each candidate heuristically
against the Ideal Candidate Profile (ICP), generates personalised cold emails
for qualified leads, and writes results to a real .xlsx Excel file.

ICP:
  Recent 2025/2026 Civil, Mechanical, Structural, Project/Construction Management
  Engineering graduates in the US, ideally with AutoCAD exposure, open to
  pivoting into Telecom/OSP (Outside Plant) engineering.

Usage:
  python pipeline.py                          # default files
  python pipeline.py my_leads.csv leads.xlsx  # custom paths
"""

import os
import csv
import re
import sys
import json
import urllib.request

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
THRESHOLD = 75

WEIGHTS = {
    "education":   35,
    "grad_year":   25,
    "skills":      25,
    "location_us": 15,
}

OUTPUT_HEADERS = [
    "Candidate Name",
    "Profile Link",
    "Calculated Fit Score",
    "Justification",
    "Generated Personalized Outreach Text",
    "Status",
]

HEADER_BG   = "1F4E79"   # dark blue
HEADER_FG   = "FFFFFF"
PENDING_BG  = "FFF2CC"   # light yellow
REJECTED_BG = "FCE4D6"   # light red

# ---------------------------------------------------------------------------
# Scraping
# ---------------------------------------------------------------------------
def map_apify_profile_to_lead(profile: dict, url: str) -> dict:
    name = profile.get("fullName") or profile.get("name") or ""
    if not name and (profile.get("firstName") or profile.get("lastName")):
        name = f"{profile.get('firstName', '')} {profile.get('lastName', '')}".strip()
    if not name:
        match = re.search(r"linkedin\.com/in/([^/]+)", url, re.IGNORECASE)
        name = match.group(1).replace("-", " ").replace("_", " ").title() if match else "Scraped Candidate"

    headline = profile.get("headline") or profile.get("title") or profile.get("position") or "LinkedIn Candidate"

    loc = profile.get("location")
    location = ""
    if isinstance(loc, str):
        location = loc
    elif isinstance(loc, dict):
        location = loc.get("name") or loc.get("country") or loc.get("city") or ""
    if not location:
        location = profile.get("locationName") or profile.get("country") or profile.get("city") or "United States"

    skills_raw = profile.get("skills")
    skills_str = ""
    if isinstance(skills_raw, list):
        parsed_skills = []
        for s in skills_raw:
            if isinstance(s, str):
                parsed_skills.append(s)
            elif isinstance(s, dict):
                parsed_skills.append(s.get("name") or s.get("title") or "")
        skills_str = ", ".join([s for s in parsed_skills if s])
    elif isinstance(skills_raw, str):
        skills_str = skills_raw

    summary = profile.get("about") or profile.get("summary") or profile.get("description") or ""

    edu_raw = profile.get("education")
    education_str = ""
    grad_year = "2025"
    if isinstance(edu_raw, list) and len(edu_raw) > 0:
        edu_list = []
        for edu in edu_raw:
            school = edu.get("schoolName") or edu.get("school") or ""
            degree = edu.get("degreeName") or edu.get("degree") or ""
            field = edu.get("fieldOfStudy") or edu.get("field") or ""

            date_val = edu.get("end") or edu.get("endDate") or edu.get("dateRange")
            edu_year = ""
            if isinstance(date_val, dict) and date_val.get("year"):
                edu_year = str(date_val.get("year"))
            elif isinstance(date_val, str):
                m = re.search(r"\b(202\d)\b", date_val)
                if m:
                    edu_year = m.group(1)
            elif isinstance(date_val, int):
                edu_year = str(date_val)

            if edu_year and 2020 <= int(edu_year) <= 2028:
                grad_year = edu_year

            detail = ""
            if degree and field:
                detail = f"{degree} in {field}"
            elif degree or field:
                detail = degree or field
            edu_list.append(f"{detail} - {school}" if detail else school)

        education_str = "; ".join([e for e in edu_list if e])
    elif isinstance(edu_raw, str):
        education_str = edu_raw

    if grad_year == "2025":
        text_to_search = f"{headline} {summary} {education_str}".lower()
        m = re.search(r"\b(202[4-6])\b", text_to_search)
        if m:
            grad_year = m.group(1)

    return {
        "LinkedIn URL": url,
        "Name": name,
        "Headline": headline,
        "Education": education_str or "Engineering Candidate",
        "Graduation Year": grad_year,
        "Skills": skills_str or "AutoCAD, Civil Engineering",
        "Location": location,
        "Summary/Bio": summary
    }

def scrape_linkedin_profiles(urls: list[str]) -> list[dict]:
    api_key = os.getenv("APIFY_API_KEY")
    if not api_key:
        print("[!] APIFY_API_KEY not found. Skipping web scraping and using default mock values.")
        return []

    actor_id = os.getenv("APIFY_ACTOR_ID", "harvestapi/linkedin-profile-scraper")
    endpoint = f"https://api.apify.com/v2/acts/{actor_id}/run-sync-get-dataset?token={api_key}"
    
    payload = json.dumps({"profileUrls": urls, "urls": urls}).encode("utf-8")
    req = urllib.request.Request(endpoint, data=payload, headers={"Content-Type": "application/json"})
    
    try:
        print(f"[*] Scraping {len(urls)} profile(s) via Apify actor '{actor_id}'...")
        with urllib.request.urlopen(req) as response:
            if response.status == 200 or response.status == 201:
                data = json.loads(response.read().decode("utf-8"))
                return data if isinstance(data, list) else []
            else:
                print(f"[!] Apify API error: {response.status}")
                return []
    except Exception as e:
        print(f"[!] Error scraping profiles: {e}")
        return []

# ---------------------------------------------------------------------------
# Scoring
# ---------------------------------------------------------------------------
def qualify_candidate_heuristically(row: dict) -> tuple[int, str]:
    """Score a candidate 0-100. Returns (score, justification string)."""
    score = 0
    justifications = []

    headline     = row.get("Headline",      "").lower()
    education    = row.get("Education",     "").lower()
    grad_year    = row.get("Graduation Year", "")
    skills       = row.get("Skills",        "").lower()
    location     = row.get("Location",      "").lower()
    bio          = row.get("Summary/Bio",   "").lower()

    combined = f"{headline} {education} {skills} {bio}"

    # 1. Education (35 pts)
    degree_map = {
        "civil":                 "Civil Engineering",
        "mechanical":            "Mechanical Engineering",
        "structural":            "Structural Engineering",
        "construction management": "Construction Management",
        "project management":    "Project Management",
    }
    is_target = False
    matched   = []
    for kw, label in degree_map.items():
        if kw in combined:
            is_target = True
            matched.append(label)

    if is_target:
        score += WEIGHTS["education"]
        justifications.append(f"Matching degree field detected: {', '.join(matched)}")
    elif "engineering" in combined or "engineer" in combined:
        score += 15
        justifications.append("Non-target engineering degree detected")
    else:
        justifications.append("Degree/field does not align with core technical ICP targets")

    # 2. Graduation year (25 pts)
    if grad_year in ("2025", "2026"):
        score += WEIGHTS["grad_year"]
        justifications.append(f"Target graduation year matched: {grad_year}")
    elif grad_year == "2024":
        score += 10
        justifications.append("Graduated 2024 — recent but slightly older than prime 2025/2026 target")
    else:
        justifications.append(
            f"Graduation year ({grad_year or 'Unknown'}) is outside the 2025-2026 target window"
        )

    # 3. AutoCAD / CAD skills (25 pts)
    if "autocad" in combined or "cad" in combined:
        score += WEIGHTS["skills"]
        justifications.append("AutoCAD/CAD modeling exposure verified")
    else:
        justifications.append("No AutoCAD or drafting software exposure detected")

    # 4. US location (15 pts)  — word-boundary regex only, no substring fallback
    us_terms = [
        "usa", "united states", r"\bus\b",
        r"\btx\b", r"\bil\b", r"\bca\b", r"\bga\b", r"\bma\b", r"\bny\b",
        "texas", "california", "georgia", "florida", "illinois",
    ]
    is_us = any(re.search(term, location) for term in us_terms)

    if is_us:
        score += WEIGHTS["location_us"]
        justifications.append("Located in the United States (eligible for immediate OSP positions)")
    else:
        justifications.append("Location is outside the US or undetermined")

    # 5. CS/Software penalty
    if ("computer science" in combined or "software" in combined) and not is_target:
        score = max(5, score - 30)
        justifications.append(
            "Profile heavily oriented towards Software/CS rather than Infrastructure Engineering"
        )

    return min(100, max(0, score)), " | ".join(justifications)


# ---------------------------------------------------------------------------
# Email generation
# ---------------------------------------------------------------------------
def generate_cold_outreach(row: dict, score: int) -> str:
    if score < THRESHOLD:
        return f"N/A - Candidate score below threshold of {THRESHOLD}"

    name      = row.get("Name", "there")
    education = row.get("Education", "your engineering studies")
    skills    = row.get("Skills", "")
    bio       = row.get("Summary/Bio", "")

    # Parse school / degree from "Degree - School" format
    school, degree = "your university", education
    if " - " in education:
        parts  = education.split(" - ", 1)
        degree = parts[0].strip()
        school = parts[1].strip()

    autocad_mention = (
        "your hands-on AutoCAD experience"
        if "autocad" in skills.lower()
        else "your drafting and design background"
    )

    # Pull one specific detail from bio to personalise
    bio_hook = ""
    if bio:
        first_sentence = bio.split(".")[0].strip()
        if len(first_sentence) > 20:
            bio_hook = f"\n\nI noticed you mentioned: \"{first_sentence.capitalize()}.\" That kind of initiative is exactly what our OSP teams look for.\n"

    return (
        f"Subject: Entry-Level OSP Engineering Opportunity — {name}\n\n"
        f"Hi {name},\n\n"
        f"I came across your profile and was impressed by your background in {degree} from {school}.{bio_hook}\n"
        f"We are actively hiring entry-level Civil, Mechanical, and Construction Engineering graduates "
        f"who are open to building a career in Telecom / Outside Plant (OSP) Engineering — "
        f"a fast-growing infrastructure sector with strong career progression.\n\n"
        f"Given {autocad_mention}, I believe your design foundation would translate directly to "
        f"mapping and detailing fiber/telecom layouts in the field.\n\n"
        f"Would you be open to a brief 10-minute call next Tuesday or Wednesday to explore if this "
        f"is a mutual fit?\n\n"
        f"Best regards,\n"
        f"Outreach Coordination Team\n"
        f"Automated Talent Acquisition Pipeline"
    )


# ---------------------------------------------------------------------------
# Excel writer — appends new rows, never wipes existing data
# ---------------------------------------------------------------------------
def _header_style() -> tuple:
    font  = Font(bold=True, color=HEADER_FG, size=11)
    fill  = PatternFill("solid", fgColor=HEADER_BG)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin  = Side(border_style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return font, fill, align, border


def _row_style(status: str) -> tuple:
    bg    = PENDING_BG if status == "Pending" else (REJECTED_BG if status == "Rejected" else "FFFFFF")
    fill  = PatternFill("solid", fgColor=bg)
    align = Alignment(vertical="top", wrap_text=True)
    thin  = Side(border_style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return fill, align, border


def write_to_excel(rows: list[dict], output_path: str) -> None:
    """Append new rows to existing .xlsx or create it fresh."""
    if os.path.exists(output_path):
        wb = load_workbook(output_path)
        ws = wb.active
        existing_links = {
            ws.cell(row=r, column=2).value
            for r in range(2, ws.max_row + 1)
        }
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Qualified Leads"
        existing_links = set()

        # Write header row
        font, fill, align, border = _header_style()
        for col_idx, header in enumerate(OUTPUT_HEADERS, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font   = font
            cell.fill   = fill
            cell.alignment = align
            cell.border = border
        ws.row_dimensions[1].height = 30

    # Append only new leads (deduplicate by Profile Link)
    new_count = 0
    for row_data in rows:
        link = row_data.get("Profile Link", "")
        if link in existing_links:
            continue

        next_row = ws.max_row + 1
        status   = row_data.get("Status", "Pending")
        fill, align, border = _row_style(status)

        for col_idx, header in enumerate(OUTPUT_HEADERS, 1):
            cell = ws.cell(row=next_row, column=col_idx, value=row_data.get(header, ""))
            cell.fill      = fill
            cell.alignment = align
            cell.border    = border
            if header == "Calculated Fit Score":
                cell.alignment = Alignment(horizontal="center", vertical="top")

        ws.row_dimensions[next_row].height = 80
        existing_links.add(link)
        new_count += 1

    # Column widths
    col_widths = [28, 45, 10, 60, 80, 12]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(output_path)
    print(f"[*] Excel file saved: {output_path}  ({new_count} new row(s) appended)")


# ---------------------------------------------------------------------------
# Main pipeline
# ---------------------------------------------------------------------------
def run_pipeline(input_csv_path: str, output_xlsx_path: str) -> None:
    print("[*] Starting Automated Lead Qualification Pipeline...")
    print(f"[*] Input  : {input_csv_path}")
    print(f"[*] Output : {output_xlsx_path}")

    if not os.path.exists(input_csv_path):
        print(f"[!] Error: Input file '{input_csv_path}' not found.")
        sys.exit(1)

    rows_to_write  = []
    qualified_count = 0

    with open(input_csv_path, mode="r", encoding="utf-8-sig") as infile:
        reader = csv.DictReader(infile)
        rows   = list(reader)

    # Scrape missing profile data for provided URLs
    urls_to_scrape = []
    for row in rows:
        url = row.get("LinkedIn URL") or row.get("Profile Link") or ""
        name = row.get("Name") or ""
        if url and (not name or name == "Scraped Candidate" or not row.get("Headline")):
            urls_to_scrape.append(url)

    if urls_to_scrape:
        scraped_data = scrape_linkedin_profiles(urls_to_scrape)
        
        # Merge scraped data into rows
        for i, row in enumerate(rows):
            url = row.get("LinkedIn URL") or row.get("Profile Link") or ""
            if url in urls_to_scrape:
                # Find matching scraped profile
                match = next((item for item in scraped_data if item.get("url", "") == url or item.get("linkedinUrl", "") == url or item.get("inputUrl", "") == url), None)
                if match:
                    mapped = map_apify_profile_to_lead(match, url)
                else:
                    mapped = map_apify_profile_to_lead({}, url) # Fallback
                
                rows[i].update(mapped)

    for idx, row in enumerate(rows, 1):
        name        = row.get("Name", f"Candidate #{idx}")
        profile_url = row.get("LinkedIn URL", "")

        print(f"[{idx}/{len(rows)}] Processing: {name}  ({profile_url})")

        score, justification = qualify_candidate_heuristically(row)
        outreach             = generate_cold_outreach(row, score)
        status               = "Pending" if score >= THRESHOLD else "Rejected"

        if status == "Pending":
            qualified_count += 1

        print(f"         Score: {score}/100  |  Status: {status}")

        rows_to_write.append({
            "Candidate Name":                    name,
            "Profile Link":                      profile_url,
            "Calculated Fit Score":              score,
            "Justification":                     justification,
            "Generated Personalized Outreach Text": outreach,
            "Status":                            status,
        })

    write_to_excel(rows_to_write, output_xlsx_path)

    print("\n" + "=" * 50)
    print("Pipeline Complete!")
    print(f"  Total processed : {len(rows)}")
    print(f"  Qualified (≥{THRESHOLD}) : {qualified_count}")
    print(f"  Rejected        : {len(rows) - qualified_count}")
    print("=" * 50)


if __name__ == "__main__":
    input_file  = sys.argv[1] if len(sys.argv) > 1 else "input_leads.csv"
    output_file = sys.argv[2] if len(sys.argv) > 2 else "output_leads.xlsx"
    run_pipeline(input_file, output_file)